import openpyxl
import os
from datetime import datetime

# Define the download directory and the path to temp.txt
download_dir = r"C:\Users\HVALDES1\PycharmProjects\SWAD_01"
temp_file_path = os.path.join(download_dir, 'temp.txt')

# Initialize an empty list to store the file paths from temp.txt
input_filepath = []

# Read the input file paths from temp.txt
with open(temp_file_path, 'r', encoding='utf-8') as file:
    for line in file:
        line = line.strip()
        if line:
            input_filepath.append(line)

print("The list of input files read from temp.txt:")
print(input_filepath)

# Path to save the concatenated output
output_filepath = 'output.xlsx'

# The name of the column to search
column_name = "Planned SW Del."  # Updated to the desired column name

# Define the months to include in the output (as strings in "MMM-YYYY" format)
months_to_include = ["Sep-2024", "Oct-2024", "Nov-2024"]  # Replace with your list of months

# Initialize the output workbook and sheet
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# Set a flag to know if the header has been written
header_written = False

# Create a set to keep track of unique rows (using string representation of rows)
seen_rows = set()

# Loop over each input file
for i in range(len(input_filepath)):
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(input_filepath[i], data_only=True)
        sheet = workbook.active  # Get the active sheet
        print(f"Processing file: {input_filepath[i]}")

        # Find the column index based on the column name
        column_index = None
        for cell in sheet[1]:  # Header row is row 1
            if cell.value == column_name:
                column_index = cell.column
                break

        if column_index is None:
            raise ValueError(f"Column '{column_name}' not found in the Excel file '{input_filepath[i]}'.")

        # Extract rows where the date falls within the specified months
        filtered_rows = []
        for row in sheet.iter_rows(min_row=2):  # Start from row 2 to skip the header
            cell_value = row[column_index - 1].value  # Adjust index for 0-based indexing

            if cell_value is not None:
                # Attempt to parse the cell value as a date
                if isinstance(cell_value, datetime):
                    date_value = cell_value
                else:
                    try:
                        # If the cell value is a string, parse it into a datetime object
                        date_value = datetime.strptime(cell_value.strip(), "%d-%b-%Y")
                    except ValueError:
                        print(f"Invalid date format in row {row[0].row}: '{cell_value}'")
                        continue  # Skip this row if the date format is incorrect

                # Extract the month and year in "MMM-YYYY" format
                month_year = date_value.strftime("%b-%Y")

                # Check if the date's month and year are in the list
                if month_year in months_to_include:
                    # Collect the cell values for the entire row
                    row_data = [cell.value for cell in row]

                    # Convert the row data to a tuple to make it hashable
                    row_tuple = tuple(row_data)

                    # Check if this row has already been seen
                    if row_tuple not in seen_rows:
                        # Add the row to the set of seen rows
                        seen_rows.add(row_tuple)
                        # Append the unique row to filtered_rows
                        filtered_rows.append(row_data)
                    else:
                        print(f"Duplicate row found and skipped: {row_data}")

        # Write the header row once
        if not header_written:
            header_row = [cell.value for cell in sheet[1]]
            new_sheet.append(header_row)
            header_written = True

        # Write the extracted unique rows to the new sheet
        for row_data in filtered_rows:
            new_sheet.append(row_data)

        print(f"Added {len(filtered_rows)} unique rows from '{input_filepath[i]}'.")

    except FileNotFoundError:
        print(f"Error: Input file '{input_filepath[i]}' not found.")
    except ValueError as e:
        print(f"Error processing '{input_filepath[i]}': {e}")
    except Exception as e:
        print(f"An unexpected error occurred with '{input_filepath[i]}': {e}")

# Save the new workbook after all files have been processed
new_workbook.save(output_filepath)
print(f"All matching unique rows have been saved to '{output_filepath}'.")
