import pandas
import os
import openpyxl
from openpyxl.utils import get_column_letter
download_dir = r"C:\Users\HVALDES1\PycharmProjects\SWAD_01"
temp_file_path = os.path.join(download_dir, 'temp.txt')
target_values = {"yes", "open"}

def extract_open_rows(input_filepath, output_filepath, column_name="6 Eyes RA"):
    """
    Extracts rows from an Excel file where the specified column contains "Open" or "Yes".

    Args:
        input_filepath: Path to the input .xlsx file.
        output_filepath: Path to save the output .xlsx file.
        column_name: The name of the column to search (default: "6 Eyes RA").
    """
    try:
        for i in range(len(string_list)):
            workbook = openpyxl.load_workbook(input_filepath[i], data_only=True)  #data_only=True to get values, not formulas
            sheet = workbook.active  # Get the active sheet
            print("input_filepath:")
            print(input_filepath[i])
            # Find the column index
            column_index = None
            for cell in sheet[1]:  # Check the header row
                if cell.value == column_name:
                    column_index = cell.column
                    break

            if column_index is None:
                raise ValueError(f"Column '{column_name}' not found in the Excel file.")

            # Extract rows
            open_rows = []

            for row in sheet.iter_rows(min_row=2):  # Start from row 2 (skip header)
                cell_value = row[column_index - 1].value  # Adjust index for 0-based indexing
                if cell_value is not None:
                    cell_value_clean = str(cell_value).strip().lower()
                    if cell_value_clean in target_values:
                        open_rows.append([cell.value for cell in row])

            # Create a new workbook and sheet
            new_workbook = openpyxl.Workbook()
            new_sheet = new_workbook.active

            # Write header row (optional)
            header_row = sheet[1]
            new_sheet.append([cell.value for cell in header_row])

            # Write extracted rows
            for row_data in open_rows:
                new_sheet.append(row_data)

            # Save the new workbook
            new_workbook.save(output_filepath)
            print(f"Rows with 'Open' or 'Yes' in '{column_name}' column saved to {output_filepath}")

    except FileNotFoundError:
        print(f"Error: Input file '{input_filepath[1]}' not found.")
    except ValueError as e:
        print(f"Error: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

# Initialize an empty list to store the strings
string_list = []

# Open the temp.txt file in read mode
with open(temp_file_path, 'r', encoding='utf-8') as file:
    # Iterate over each line in the file
    for line in file:
        # Remove any leading/trailing whitespace and newline characters
        line = line.strip()
        # Append the line to the list if it's not empty
        if line:
            string_list.append(line)

# Print the list to verify the contents
print("The list of strings read from temp.txt:")
print(string_list)

output_file = "output.xlsx" #Path to save the output .xlsx file

extract_open_rows(string_list, output_file)